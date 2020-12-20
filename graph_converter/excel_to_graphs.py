from openpyxl import load_workbook, Workbook

EXCEL_EXPORT_FILE_PATH = './graph_export.xlsx'
HEROES_LIST_FILE_PATH = '../src/main/resources/heroes_list'
ENEMY_GRAPH_FILE_PATH = '../src/main/resources/enemy_graph'
TEAM_GRAPH_FILE_PATH = '../src/main/resources/team_graph'
VALUE_SEPARATOR = ','


def export_heroes_list(wb: Workbook):
    ws = wb.get_sheet_by_name('Heroes List')
    nb_heroes = ws.max_row
    with open(HEROES_LIST_FILE_PATH, 'w', encoding='UTF-8') as f:
        for i, (name_cell, role_cell) in enumerate(ws['A1':'B'+str(nb_heroes)]):
            if i > 0:
                f.write('\n')
            f.write(f'{name_cell.value}{VALUE_SEPARATOR}{role_cell.value}')


def export_graph(wb: Workbook, sheet_name: str, file_path: str):
    ws = wb.get_sheet_by_name(sheet_name)
    nb_rows = ws.max_row - 1  # ignore header
    with open(file_path, 'w', encoding='UTF-8') as f:
        for max_col, row in enumerate(ws['2':str(nb_rows+1)]):
            if max_col > 0:
                f.write('\n')
            f.write(VALUE_SEPARATOR.join(str(row[i+1].value) for i in range(0, max_col + 1)))


def main():
    wb = load_workbook(EXCEL_EXPORT_FILE_PATH, read_only=True)
    export_heroes_list(wb)
    export_graph(wb, 'Synergies', TEAM_GRAPH_FILE_PATH)
    export_graph(wb, 'Counters', ENEMY_GRAPH_FILE_PATH)


if __name__ == '__main__':
    main()
