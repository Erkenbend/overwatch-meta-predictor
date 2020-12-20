from openpyxl import Workbook
from openpyxl.styles import Font

MODIFIABLE_CELLS_FONT = Font(bold=True)
REF_CELLS_FONT = Font(color='969696')
EXCEL_EXPORT_FILE_PATH = 'graph_export.xlsx'
HEROES_LIST_FILE_PATH = '../src/main/resources/heroes_list'
ENEMY_GRAPH_FILE_PATH = '../src/main/resources/enemy_graph'
TEAM_GRAPH_FILE_PATH = '../src/main/resources/team_graph'
VALUE_SEPARATOR = ','


def load_heroes_list(path: str):
    """
    read config file and return heroes as tuple of tuples:
    ((Tracer, D), (Sigma, T), ...)
    """
    heroes_list = []
    with open(path, 'r', encoding='UTF-8') as f:
        lines = f.read().split('\n')
        heroes_list.extend(tuple(line.split(VALUE_SEPARATOR)) for line in lines)
    return tuple(heroes_list)


def load_graph(path: str):
    """
    read graph file and return coefficients as tuple of tuples:
    ((1), (1, 0), (-1, 1, 0), ...)
    """
    graph = []
    with open(path, 'r', encoding='UTF-8') as f:
        lines = f.read().split('\n')
        graph.extend(tuple(map(int, line.split(VALUE_SEPARATOR))) for line in lines)
    return tuple(graph)


def get_synergy(pos_hero_1, pos_hero_2, team_graph):
    """
    return synergy between 2 heroes given team graph
    """
    if pos_hero_2 > pos_hero_1:
        return team_graph[pos_hero_2][pos_hero_1]
    return team_graph[pos_hero_1][pos_hero_2]


def get_dominance(pos_hero_1, pos_hero_2, enemy_graph):
    """
    return dominance from on hero over the other given enemy graph
    """
    if pos_hero_2 > pos_hero_1:
        return -enemy_graph[pos_hero_2][pos_hero_1]
    return enemy_graph[pos_hero_1][pos_hero_2]


def initialize_workbook() -> Workbook:
    wb = Workbook()
    return wb


def create_heroes_list_worksheet(wb: Workbook, heroes_list):
    ws = wb.create_sheet("Heroes List")
    wb.active = ws

    for i, hero in enumerate(heroes_list):
        name_cell = ws['A'+str(i+1)]
        name_cell.value = hero[0]
        name_cell.font = MODIFIABLE_CELLS_FONT
        role_cell = ws['B'+str(i+1)]
        role_cell.value = hero[1]
        role_cell.font = MODIFIABLE_CELLS_FONT


def remove_default_worksheet(wb: Workbook):
    default_ws = wb.get_sheet_by_name('Sheet')
    wb.remove(default_ws)


def create_headers(ws, heroes_list):
    for i, hero in enumerate(heroes_list):
        current_hero_name_ref = f'=\'Heroes List\'!A{str(i + 1)}'
        name_cell_left_col = ws['A' + str(i + 2)]
        name_cell_left_col.value = current_hero_name_ref
        name_cell_left_col.font = REF_CELLS_FONT
        name_cell_top_row = ws[col_as_char(i + 2) + '1']
        name_cell_top_row.value = current_hero_name_ref
        name_cell_top_row.font = REF_CELLS_FONT


def create_synergy_matrix(wb: Workbook, heroes_list, team_graph):
    ws = wb.create_sheet("Synergies")
    create_headers(ws, heroes_list)
    for i in range(len(heroes_list)):
        # synergy with oneself is 0 per symmetry
        self_synergy_cell = ws[col_as_char(i+2)+str(i+2)]
        self_synergy_cell.value = 0
        self_synergy_cell.font = REF_CELLS_FONT
        for j in range(i):
            # fill lower-left side of the matrix
            modifiable_cell = ws[col_as_char(j+2)+str(i+2)]
            modifiable_cell.value = get_synergy(i, j, team_graph)
            modifiable_cell.font = MODIFIABLE_CELLS_FONT
            # use refs for upper-right side
            ref_cell = ws[col_as_char(i+2)+str(j+2)]
            ref_cell.value = f'={col_as_char(j+2)+str(i+2)}'
            ref_cell.font = REF_CELLS_FONT


def create_dominance_matrix(wb: Workbook, heroes_list, enemy_graph):
    ws = wb.create_sheet("Counters")
    create_headers(ws, heroes_list)
    for i in range(len(heroes_list)):
        # dominance over oneself is 0 per symmetry
        self_dominance_cell = ws[col_as_char(i+2)+str(i+2)]
        self_dominance_cell.value = 0
        self_dominance_cell.font = REF_CELLS_FONT
        for j in range(i):
            # fill lower-left side of the matrix
            modifiable_cell = ws[col_as_char(j+2)+str(i+2)]
            modifiable_cell.value = get_dominance(i, j, enemy_graph)
            modifiable_cell.font = MODIFIABLE_CELLS_FONT
            # use refs for upper-right side
            ref_cell = ws[col_as_char(i+2)+str(j+2)]
            ref_cell.value = f'=-{col_as_char(j+2)+str(i+2)}'
            ref_cell.font = REF_CELLS_FONT


def col_as_char(i: int):
    """
    helper function to convert int into letter of excel col nb
    1 --> A
    2 --> B
    26 --> Z
    27 --> AA
    28 --> AB
    29 --> AC
    """
    nb_letters = 26
    # handle special case that happens when calculating modulo
    if i == 0:
        i = nb_letters
    # find smallest power of 26 that is big enough to write number with letters
    current_pow, current_max = 1, nb_letters
    while i > current_max:
        current_pow += 1
        current_max *= (nb_letters+1)
    # recursively handle extra letters
    if current_pow == 1:
        return chr(ord('A')+i-1)
    return col_as_char((i-1) // nb_letters) + col_as_char(i % nb_letters)


def main():
    heroes_list = load_heroes_list(HEROES_LIST_FILE_PATH)
    # print(heroes_list)
    team_graph = load_graph(TEAM_GRAPH_FILE_PATH)
    # print(team_graph)
    enemy_graph = load_graph(ENEMY_GRAPH_FILE_PATH)
    # print(enemy_graph)

    wb = initialize_workbook()

    create_heroes_list_worksheet(wb, heroes_list)
    remove_default_worksheet(wb)
    create_synergy_matrix(wb, heroes_list, team_graph)
    create_dominance_matrix(wb, heroes_list, enemy_graph)

    try:
        wb.save(EXCEL_EXPORT_FILE_PATH)
    except PermissionError:
        print('Could not save, please check that file is closed.')
        exit(1)


if __name__ == '__main__':
    main()
